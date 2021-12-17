# tkinter is a module used to product the graphical user interface (GUI)
import tkinter as tk
from tkinter import ttk
# messagebox allows pop-up windows in the application
from tkinter import messagebox
# simpledialog is a messagebox that allows for input
from tkinter import simpledialog
# pandas and numpy are data analysis packages; the two most-common packages in Python, likely
import pandas as pd
import numpy as np
# openpyxl allows for reading/writing from/to Excel files, rather than CSV, which restricts formatting options
from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
# Very standard math module, used only for the floor function
from math import floor
# datetime module for converting strings to dates, and calculating the difference between times entered in the fields
from datetime import datetime, timedelta
# For accessing the text logs folder
import os
from os import listdir  # , getcwd
from os.path import isfile, join

CURRENT_YEAR = str(datetime.now().year)
MONTH_ENTERED = ''

PATH_MONTHLY = ''
PATH_MASTER = ''
PATH_MASTER_COPY = ''
PATH_DRAFTS = ''
PATH_LOGS = ''
PATH_IMAGE = ''

MAX_TEXT_FILES = 50


class App(tk.Tk):

    def __init__(self):
        '''The class initialization function first instantiates some data structures that are referenced later, like the list of
           service call types. The errors dictionary below tracks the data validity when the Submit button is clicked. After
           instantiation, functions are called that create the GUI elements, like labels, entry boxes, and buttons. Then, window
           settings are set, and the image for the application icon is imported.'''

        tk.Tk.__init__(self)

        self.service_call_type_list = ['NONE', 'Access', 'Alarm', 'Arrest', 'Assist Police/EMS', 'By Law', 'Camera Audit',
                                       'Camera Footage Review', 'Camera Malfunction', 'Code Red', 'Daily Lock/Unlock',
                                       'Elevator Kirkwood', 'Emergency Card Swipe Testing', 'Escort Delivery', 'Evidence/Contraband',
                                       'Facility Maintenance', 'Fall No Injuries', 'Fall Unknown Injuries', 'Fall With Injuries',
                                       'Guard Duties - Other', 'Information', 'Lock/Unlock Door', 'Monitor Camera',
                                       'Motor Vehicle Accident', 'Off-Site Checks', 'Off-Site Checks Cancelled',
                                       'Off-Site Service Calls', 'One-to-One', 'Other Service Calls', 'Parking',
                                       'Patrol Duties', 'POI', 'Search Room', 'Side Room Entry', 'Staff Falls',
                                       'Visitor - Security Presence/Assistance', 'Weekly Audits']

        self.errors = {'date': 1,
                       'call_received': 1,
                       'arrival_time': 1,
                       'completion_time': 1,
                       'service_call_type': 1, }

        self.months = {'1': '01 - January', '2': '02 - February', '3': '03 - March', '4': '04 - April', '5': '05 - May', '6': '06 - June',
                       '7': '07 - July', '8': '08 - August', '9': '09 - September', '10': '10 - October', '11': '11 - November', '12': '12 - December'}

        self.handle_label_creation()
        self.handle_radio_button_creation()
        self.handle_entry_creation()
        self.handle_listbox_creation()
        self.handle_checkbox_creation()
        self.handle_second_entry_creation()
        self.handle_textbox_creation()
        self.handle_second_checkbox_creation()
        self.handle_button_creation()

        # Index of draft, if opened. If submitted, draft will be deleted
        self.draft_opened_index = None

        # Window settings
        self.resizable(False, False)
        self.winfo_toplevel().title('Incident Entry Tool')
        self.window = None  # This is to check later if a toplevel window already exists

        try:
            self.iconbitmap(PATH_IMAGE)
        except:
            pass

    def get_dataframe(self):
        '''Load the monthly dataframe. If it doesn't exist, create an empty one.'''

        try:
            self.df = pd.read_excel(PATH_MONTHLY + CURRENT_YEAR + '\\' + MONTH_ENTERED +
                                    '\\Incident Reports - ' + CURRENT_YEAR + ' ' + MONTH_ENTERED[5:] + '.xlsx')

        except:
            self.df = pd.DataFrame(columns=['Date', 'Time Entered', 'Shift', 'Call Received Time', 'Arrival Time', 'Completion Time', 'Service Call Type',
                                            'Physical Intervention', 'Restraint Used', 'Police Involved', 'Requested By', 'Contact Information',
                                            'Notes', 'Time Taken to Arrive', 'Time Taken From Call to Completion', 'Time Taken From Arrival to Completion',
                                            'Time Taken to Arrive (mins.)', 'Time Taken From Call to Completion (mins.)', 'Time Taken From Arrival to Completion (mins.)'])

    def get_master_dataframe(self):
        '''Load the master dataframe. If it doesn't exist, DON'T create an empty one...display a error message.'''

        try:
            self.master_df = pd.read_excel(PATH_MASTER)

        except:
            tk.messagebox.showinfo(
                'Data Load Error', 'The Master file can not be found.')

    def get_saves_dataframe(self):
        '''Load the drafts dataframe. If it doesn't exist, create an empty one. Values aren't validated, so integers can come in as
           floats. Use the converters argument to read_excel to turn the integers into strings.'''

        try:
            self.saves_df = pd.read_excel(PATH_DRAFTS, converters={
                                          0: str, 2: str, 3: str, 4: str, 9: str, 10: str, 11: str})

        except:
            self.saves_df = pd.DataFrame(columns=['Identifier', 'Date', 'Shift', 'Call Received Time', 'Arrival Time', 'Completion Time', 'Service Call Type',
                                                  'Physical Intervention', 'Restraint Used', 'Police Involved', 'Requested By', 'Contact Information',
                                                  'Notes', 'Time Over 24 Hours'])

    def handle_label_creation(self):
        '''Create the text labels that accompany the widgets. Also, create some horizontal lines for aesthetics and spacing.'''

        self.shift_label = tk.Label(self, font=('Calibri', 12), text='SHIFT:').grid(
            row=0, column=0, pady=(10, 0))
        self.date_label = tk.Label(self, font=(
            'Calibri', 12), text='Date (mm/dd):').grid(row=2, column=0, sticky='e')
        self.call_received_time_label = tk.Label(self, font=(
            'Calibri', 12), text='Call Received (24hr.):').grid(row=3, column=0, sticky='e')
        self.arrival_time_label = tk.Label(self, font=(
            'Calibri', 12), text='Arrival Time (24hr.):').grid(row=4, column=0, sticky='e')
        self.completion_time_label = tk.Label(self, font=(
            'Calibri', 12), text='Completion Time (24hr.):').grid(row=5, column=0, sticky='e', padx=(10, 0))
        self.service_call_type_label = tk.Label(self, font=(
            'Calibri', 12), text='Service Call Type:').grid(row=7, column=0, padx=(17, 0))
        self.physical_intervention_label = tk.Label(self, font=(
            'Calibri', 12), text='Physical Intervention:').grid(row=9, column=0, sticky='e')
        self.restraint_used_label = tk.Label(self, font=(
            'Calibri', 12), text='Restraint Used:').grid(row=10, column=0, sticky='e')
        self.police_involved_label = tk.Label(self, font=(
            'Calibri', 12), text='Police Involved:').grid(row=11, column=0, sticky='e')
        self.requested_by_label = tk.Label(self, font=('Calibri', 12), text='Requested By:').grid(
            row=12, column=0, sticky='e', pady=(15, 0))
        self.contact_information_label = tk.Label(self, font=(
            'Calibri', 12), text='Contact Information:').grid(row=13, column=0, sticky='e')
        self.notes_label = tk.Label(self, font=('Calibri', 14), text='Notes:').grid(
            columnspan=2, row=14, pady=(15, 3))
        self.completion_time = tk.Label(self, font=('Calibri', 11), text='Was Time From Call to Completion >24 Hours?').grid(
            columnspan=2, row=17, column=0, sticky='w', padx=(20, 0), pady=(5, 3))

        # Horizontal Lines
        self.h_line = ttk.Separator(self, orient='horizontal')
        self.h_line_2 = ttk.Separator(self, orient='horizontal')
        self.h_line_3 = ttk.Separator(self, orient='horizontal')

        self.h_line.grid(row=1, column=0, sticky='we',
                         columnspan=2, pady=10, padx=12)
        self.h_line_2.grid(row=6, column=0, sticky='we',
                           columnspan=2, pady=10, padx=12)
        self.h_line_3.grid(row=18, column=0, sticky='we',
                           columnspan=2, pady=10, padx=12)

    def handle_radio_button_creation(self):
        '''Create the single set of radio buttons for selecting the shift.'''

        self.var = tk.StringVar()
        self.var.set('7:30 - 19:30')

        self.date_morning_radio = tk.Radiobutton(
            self, text='7:30 - 19:30', padx=20, variable=self.var, value='7:30 - 19:30')
        self.date_morning_radio.grid(row=0, column=1, sticky='w', pady=(10, 0))

        self.date_night_radio = tk.Radiobutton(
            self, text='19:30 - 7:30', padx=20, variable=self.var, value='19:30 - 7:30')
        self.date_night_radio.grid(
            row=0, column=1, padx=(110, 0), pady=(10, 0))

    def handle_entry_creation(self):
        '''Create the first set of entry widgets, for date and the three times.'''

        self.date_entry = tk.Entry(self, width=5)
        self.date_entry.bind('<FocusOut>', self.focus_date_validation)
        self.date_entry.grid(row=2, column=1)

        self.call_received_entry = tk.Entry(self, width=5)
        self.call_received_entry.bind(
            '<FocusOut>', self.focus_call_received_validation)
        self.call_received_entry.grid(row=3, column=1)

        self.arrival_time_entry = tk.Entry(self, width=5)
        self.arrival_time_entry.bind(
            '<FocusOut>', self.focus_arrival_time_validation)
        self.arrival_time_entry.grid(row=4, column=1)

        self.completion_time_entry = tk.Entry(self, width=5)
        self.completion_time_entry.bind(
            '<FocusOut>', self.focus_completion_time_validation)
        self.completion_time_entry.grid(row=5, column=1)

    def handle_second_entry_creation(self):
        '''Create the two non-validated entry fields for Requested By and Contact Information'''

        self.requested_by_entry = tk.Entry(self, width=30)
        self.requested_by_entry.grid(row=12, column=1, pady=(20, 0))

        self.contact_information_entry = tk.Entry(self, width=30)
        self.contact_information_entry.grid(row=13, column=1)

    def handle_checkbox_creation(self):
        '''Create the checkboxes for Physical Intervention, Restraint Used, and Police Involved.'''

        self.physical_intervention_checkbox = ttk.Checkbutton(self, text='Yes')
        self.physical_intervention_checkbox.grid(row=9, column=1)
        self.physical_intervention_checkbox.state(['!alternate'])

        self.restraint_used_checkbox = ttk.Checkbutton(self, text='Yes')
        self.restraint_used_checkbox.grid(row=10, column=1)
        self.restraint_used_checkbox.state(['!alternate'])

        self.police_involved_checkbox = ttk.Checkbutton(self, text='Yes')
        self.police_involved_checkbox.grid(row=11, column=1)
        self.police_involved_checkbox.state(['!alternate'])

    def handle_second_checkbox_creation(self):
        '''Create the checkbox for the Over 24 Hour indicator.'''

        self.time_over_24_hours_checkbox = ttk.Checkbutton(self, text='Yes')
        self.time_over_24_hours_checkbox.grid(
            columnspan=2, row=17, column=0, sticky='e', padx=(0, 45), pady=(5, 2))
        self.time_over_24_hours_checkbox.state(['!alternate'])

    def handle_textbox_creation(self):
        '''Create the textbox for Notes.'''

        self.notes_textbox = tk.Text(
            self, font=('Calibri', 9), height=3, width=60)
        self.notes_textbox.grid(columnspan=2, row=15, padx=10)

    def handle_button_creation(self):
        '''Create the Submit, Save Draft, and Load Saved Entry buttons.'''

        self.submit_button = tk.Button(self, font=(
            'Arial', 16, 'bold'), text="SUBMIT", command=self.on_submit_button)
        self.submit_button.grid(columnspan=2, row=16, pady=10)

        self.save_button = tk.Button(self, font=(
            'Arial', 11), fg='blue', text="Save Draft and Exit", command=self.on_save_button)
        self.save_button.grid(row=19, column=0, padx=(
            10, 0), pady=(0, 10), sticky='w')

        self.load_button = tk.Button(self, font=(
            'Arial', 11), fg='blue', text="Load Saved Entry", command=self.on_load_button)
        self.load_button.grid(row=19, column=1, padx=(
            0, 10), pady=(0, 10), sticky='e')

    def handle_listbox_creation(self):
        '''Create the variable for the Entry field. Bind a function to it that updates the listbox based on the search query.
           Create the Entry, and also a vertical scrollbar to use with the listbox. Bind a function to the listbox that will
           validate the contents of the Entry field when focus leaves the listbox.'''

        self.search_var = tk.StringVar()
        # Bind function to the variable in the tray name entry field
        self.search_var.trace('w', self.update_list)

        self.service_call_type_entry = tk.Entry(
            self, textvariable=self.search_var, width=30)
        self.service_call_type_entry.grid(
            row=7, column=1, padx=(5, 32))  # ENTRY WIDGET

        self.scrollbar = tk.Scrollbar(self, orient='vertical')
        self.scrollbar.grid(row=8, column=1, pady=10, sticky='nes')

        self.service_call_type_listbox = tk.Listbox(
            self, width=55, height=5, activestyle='none', yscrollcommand=self.scrollbar.set)
        self.service_call_type_listbox.bind(
            '<FocusOut>', self.focus_service_call_type_validation)
        self.service_call_type_listbox.grid(
            columnspan=2, row=8, padx=10, pady=(10, 15))

        self.scrollbar.config(
            command=self.service_call_type_listbox.yview)  # move two up

        # Function for updating the list/doing the search. It needs to be called here to populate the listbox.
        self.update_list()

    def update_list(self, *args):
        '''First, clear any background coloring of the Entry field. Get the Entry field query, and delete all listbox entries.
           For each item in the Service Call Type list, if the query is a substirng of that item, insert the item into the listbox.'''

        self.service_call_type_entry.config(
            {'background': 'White'})  # Reset the text box

        search_term = self.search_var.get()
        self.service_call_type_listbox.delete(0, 'end')

        for item in self.service_call_type_list:
            if search_term.lower().strip() in item.lower():
                self.service_call_type_listbox.insert('end', '  ' + item)

#########################################################################################################

    def focus_date_validation(self, event=None):
        '''See if the date can be parsed from the Entry field. If so, change the background color to green.
           If the date can't be parsed, if the Entry is blank, make the background color white, otherwise red.'''

        try:
            datetime.strptime(self.date_entry.get().strip(), '%m/%d')
            self.date_entry.config({'background': '#00cc2c'})
        except:
            if self.date_entry.get().strip() != '':
                self.date_entry.config({'background': 'Red'})
            else:
                self.date_entry.config({'background': 'White'})

    def focus_call_received_validation(self, event=None):
        '''If the time string length is less than 4, make the background color red. Otherwise, try to parse the time.
           It doesn't matter if a colon was used. Otherwise, if the Entry is blank, make the background color white. If
           not blank, then make the background color red.'''

        try:
            if len(self.call_received_entry.get().strip()) < 4:
                raise Exception
            datetime.strptime(
                self.call_received_entry.get().strip().replace(':', ''), '%H%M')
            self.call_received_entry.config({'background': '#00cc2c'})
        except:
            if self.call_received_entry.get().strip() != '':
                self.call_received_entry.config({'background': 'Red'})
            else:
                self.call_received_entry.config({'background': 'White'})
                self.call_received_entry.delete(0, 'end')

    def focus_arrival_time_validation(self, event=None):
        '''Same logic as above.'''

        try:
            if len(self.arrival_time_entry.get().strip()) < 4:
                fail = 1/0
            datetime.strptime(
                self.arrival_time_entry.get().strip().replace(':', ''), '%H%M')
            self.arrival_time_entry.config({'background': '#00cc2c'})
        except:
            if self.arrival_time_entry.get().strip() != '':
                self.arrival_time_entry.config({'background': 'Red'})
            else:
                self.arrival_time_entry.config({'background': 'White'})
                self.arrival_time_entry.delete(0, 'end')

    def focus_completion_time_validation(self, event=None):
        '''Same logic as above.'''

        try:
            if len(self.completion_time_entry.get().strip()) < 4:
                fail = 1/0
            datetime.strptime(
                self.completion_time_entry.get().strip().replace(':', ''), '%H%M')
            self.completion_time_entry.config({'background': '#00cc2c'})
        except:
            if self.completion_time_entry.get().strip() != '':
                self.completion_time_entry.config({'background': 'Red'})
            else:
                self.completion_time_entry.config({'background': 'White'})
                self.completion_time_entry.delete(0, 'end')

    def focus_service_call_type_validation(self, event=None):
        '''If an item in the listbox is selected, insert it into the Entry field and set the background color to green.
           Otherwise, if the entry field value matches an item in the service call type list or is blank, make the background green.
           Otherwise, make the background red.'''

        if self.service_call_type_listbox.get('active').strip() in self.service_call_type_list:
            # Because when you delete tray_entry text, the 'anchor' becomes ''
            temp_string = self.service_call_type_listbox.get('active').strip()
            self.service_call_type_entry.delete(
                0, 'end')  # This resets the 'anchor' text
            self.service_call_type_entry.insert('end', ' ' + temp_string)
            self.service_call_type_entry.config({"background": "#00cc2c"})

        elif self.service_call_type_entry.get().strip() in self.service_call_type_list:
            self.service_call_type_entry.config({"background": "#00cc2c"})

        elif self.service_call_type_entry.get().strip() == '':
            self.service_call_type_entry.config({"background": "White"})

        else:
            self.service_call_type_entry.config({"background": "Red"})

#########################################################################################################

    def date_validation(self, event=None):
        '''See if the date can be parsed from the Entry field. If so, store it in a variable, and set the month for later.
           Change the background color to green, and indicate that there is no error.
           If the date can't be parsed, set the background color to red and leave the error marked in the dictionary.'''

        try:
            datetime.strptime(self.date_entry.get().strip(), '%m/%d')
            # Make sure the date is in the proper format
            entered_date = datetime.strptime(
                self.date_entry.get().strip(), '%m/%d')
            global MONTH_ENTERED
            MONTH_ENTERED = self.months[str(entered_date.month)]

            self.date_entry.config({'background': '#00cc2c'})
            self.errors['date'] = 0
        except:
            self.date_entry.config({'background': 'Red'})
            self.errors['date'] = 1

    def call_received_validation(self, event=None):
        '''If the time string length is less than 4, make the background color red and set an error. Otherwise, try to parse the time.
           It doesn't matter if a colon was used. If valid, set the background color to green and set the errors dictionary value to 0.
           If not valid, change the background color to red and set an error.'''

        try:
            if len(self.call_received_entry.get().strip()) < 4:
                fail = 1/0
            datetime.strptime(
                self.call_received_entry.get().strip().replace(':', ''), '%H%M')
            self.call_received_entry.config({'background': '#00cc2c'})
            self.errors['call_received'] = 0
        except:
            self.call_received_entry.config({'background': 'Red'})
            self.errors['call_received'] = 1

    def arrival_time_validation(self, event=None):
        '''Same logic as above.'''

        try:
            if len(self.arrival_time_entry.get().strip()) < 4:
                fail = 1/0
            datetime.strptime(
                self.arrival_time_entry.get().strip().replace(':', ''), '%H%M')
            self.arrival_time_entry.config({'background': '#00cc2c'})
            self.errors['arrival_time'] = 0
        except:
            self.arrival_time_entry.config({'background': 'Red'})
            self.errors['arrival_time'] = 1

    def completion_time_validation(self, event=None):
        '''Same logic as above.'''

        try:
            if len(self.completion_time_entry.get().strip()) < 4:
                fail = 1/0
            datetime.strptime(
                self.completion_time_entry.get().strip().replace(':', ''), '%H%M')
            self.completion_time_entry.config({'background': '#00cc2c'})
            self.errors['completion_time'] = 0
        except:
            self.completion_time_entry.config({'background': 'Red'})
            self.errors['completion_time'] = 1

    def service_call_type_validation(self, event=None):
        '''If an item in the listbox is selected, insert it into the Entry field and set the background color to green. Clear the error.
           Otherwise, if the entry field value matches an item in the service call type list, make the background green and clear the error.
           Otherwise, make the background red and set an error.'''

        if self.service_call_type_listbox.get('active').strip() in self.service_call_type_list:
            # Because when you delete tray_entry text, the 'anchor' becomes ''
            temp_string = self.service_call_type_listbox.get('active').strip()
            self.service_call_type_entry.delete(
                0, 'end')  # This resets the 'anchor' text
            self.service_call_type_entry.insert('end', ' ' + temp_string)
            self.service_call_type_entry.config({"background": "#00cc2c"})
            self.errors['service_call_type'] = 0

        elif self.service_call_type_entry.get().strip() in self.service_call_type_list:
            self.service_call_type_entry.config({"background": "#00cc2c"})
            self.errors['service_call_type'] = 0

        else:
            self.service_call_type_entry.config({"background": "Red"})
            self.errors['service_call_type'] = 1

##########################################################################################################

    def get_checkbox_answers(self):
        '''Preset the answers to be 'No'. The Checkboxes' state returns a tuple. 'selected' can be in index 0 or 1.
           It depends on whether it is in focus. So, check at both tuple indices.'''

        self.physical_intervention_answer = 'No'
        self.restraint_used_answer = 'No'
        self.police_involved_answer = 'No'
        self.time_over_24_hours_answer = 'No'

        try:
            # Output will be one of: (), ('focus'), ('selected'), or ('focus', 'selected')
            if (self.physical_intervention_checkbox.state()[0] == 'selected' or self.physical_intervention_checkbox.state()[1] == 'selected'):
                self.physical_intervention_answer = 'Yes'
        except:
            pass

        try:
            if (self.restraint_used_checkbox.state()[0] == 'selected' or self.restraint_used_checkbox.state()[1] == 'selected'):
                self.restraint_used_answer = 'Yes'
        except:
            pass

        try:
            if (self.police_involved_checkbox.state()[0] == 'selected' or self.police_involved_checkbox.state()[1] == 'selected'):
                self.police_involved_answer = 'Yes'
        except:
            pass

        try:
            if (self.time_over_24_hours_checkbox.state()[0] == 'selected' or self.time_over_24_hours_checkbox.state()[1] == 'selected'):
                self.time_over_24_hours_answer = 'Yes'
        except:
            pass

    def format_date(self):
        '''Ensures that the entered date has a '/' inside it.'''

        month_substring, date_substring = self.date_entry.get().strip().split('/')

        if len(month_substring) == 1:
            month_substring = '0' + month_substring

        if len(date_substring) == 1:
            date_substring = '0' + date_substring

        return (CURRENT_YEAR + '/' + month_substring + '/' + date_substring)

    def format_time(self, time):
        '''Adds a colon to the entered time if not present.'''

        if ':' not in time:
            time = time[:-2] + ':' + time[-2:]

        return time

    def get_time_difference_numeric(self, time_1, time_2, unit, check_24=None):
        '''Convert the strings to datetime objects and find the difference in seconds. Based on the unit specified, calculate the
           time difference and return it.'''

        time_1 = datetime.strptime(self.format_time(time_1.strip()), '%H:%M')
        time_2 = datetime.strptime(self.format_time(time_2.strip()), '%H:%M')

        seconds = (time_2 - time_1).seconds

        if unit == 'seconds':
            if ((self.time_over_24_hours_answer == 'Yes') and (check_24 == 'Yes')):
                return str(seconds) + 86400
            else:
                return str(seconds)
        elif unit == 'minutes':
            if ((self.time_over_24_hours_answer == 'Yes') and (check_24 == 'Yes')):
                return str(round(seconds/60) + 1440)
            else:
                return str(round(seconds/60))
        elif unit == 'hours':
            if ((self.time_over_24_hours_answer == 'Yes') and (check_24 == 'Yes')):
                return str(round(seconds/3600, 1) + 24)
            else:
                return str(round(seconds/3600, 1))

    def get_time_difference(self, time_1, time_2, check_24=None):
        '''Convert the strings into Datetime objects. Subtract them to get the seconds. Calculate the minutes by dividing by 60 and
           rounding, and hours by dividing by 3600 and taking the floor. If time difference is more than 60 minutes, recalculate minutes.
           Possibly add 24 hours to the hours, create an hours string, and then return that string with minutes also. If time difference is
           not more than 60 minutes, maybe add 24 hours, and then return the hours + minutes strings.'''

        time_1 = datetime.strptime(self.format_time(time_1.strip()), '%H:%M')
        time_2 = datetime.strptime(self.format_time(time_2.strip()), '%H:%M')

        seconds = (time_2 - time_1).seconds
        minutes = str(round(seconds/60))
        hours = floor(seconds/60/60)

        hours_string = ''

        if seconds/60 > 59:
            # Minutes past the floored hour
            minutes = str(round((seconds/60/60 - hours) * 60))

            # check_24: time from call to arrival isn't increased by 24
            if ((self.time_over_24_hours_answer == 'Yes') and (check_24 == 'Yes')):
                hours = hours + 24

            hours_string = str(hours) + ' hours, '
            return (hours_string + minutes + ' minutes').replace('1 hours', '1 hour').replace('1 minutes', '1 minute').replace(', 0 minutes', '')

        elif ((self.time_over_24_hours_answer == 'Yes') and (check_24 == 'Yes')):
            hours_string = '24 hours, '

        return (hours_string + minutes + ' minutes').replace('1 hours', '1 hour').replace('1 minutes', '1 minute').replace(', 0 minutes', '')

    def append_row_to_df(self):
        '''Get validated entry values, and append to the imported monthly dataframe. Also append to the imported master dataframe.'''

        self.row_to_append = {
            'Date': self.format_date(),
            'Time Entered': str(datetime.now())[0:16],
            'Shift': self.var.get(),
            'Call Received Time': self.format_time(self.call_received_entry.get().strip()),
            'Arrival Time': self.format_time(self.arrival_time_entry.get().strip()),
            'Completion Time': self.format_time(self.completion_time_entry.get().strip()),
            'Service Call Type': self.service_call_type_entry.get().strip(),
            'Physical Intervention': self.physical_intervention_answer,
            'Restraint Used': self.restraint_used_answer,
            'Police Involved': self.police_involved_answer,
            'Requested By': self.requested_by_entry.get().strip(),
            'Contact Information': self.contact_information_entry.get().strip(),
            'Notes': self.notes_textbox.get('1.0', 'end-1c'),
            'Time Taken to Arrive': self.get_time_difference(self.call_received_entry.get(), self.arrival_time_entry.get()),
            'Time Taken From Call to Completion': self.get_time_difference(self.call_received_entry.get(), self.completion_time_entry.get(), 'Yes'),
            'Time Taken From Arrival to Completion': self.get_time_difference(self.arrival_time_entry.get(), self.completion_time_entry.get(), 'Yes'),
            'Time Taken to Arrive (mins.)': self.get_time_difference_numeric(self.call_received_entry.get(), self.arrival_time_entry.get(), 'minutes'),
            'Time Taken From Call to Completion (mins.)': self.get_time_difference_numeric(self.call_received_entry.get(), self.completion_time_entry.get(), 'minutes', 'Yes'),
            'Time Taken From Arrival to Completion (mins.)': self.get_time_difference_numeric(self.arrival_time_entry.get(), self.completion_time_entry.get(), 'minutes', 'Yes')
        }

        # THIS WILL APPEND TO THE MASTER DATAFRAME ALSO, TO SAVE DUPLICATING THIS FUNCTION
        self.master_df = self.master_df.append(
            self.row_to_append, ignore_index=True)
        # This line nor any following lines/functions will run if the Master file is not found
        self.df = self.df.append(self.row_to_append, ignore_index=True)

    def append_row_to_saves_df(self):
        '''Get the values for all the columns to be saved. If none are blank, ask for information to later identify the draft.
           If user clicked 'Ok', append to the dataframe.'''

        # Check to make sure data has been entered before saving
        self.row_to_append_saves_test = {'Date': self.date_entry.get().strip(),
                                         'Shift': self.var.get().strip(),
                                         'Call Received Time': self.call_received_entry.get().strip(),
                                         'Arrival Time': self.arrival_time_entry.get().strip(),
                                         'Completion Time': self.completion_time_entry.get().strip(),
                                         'Service Call Type': self.service_call_type_entry.get().strip(),
                                         'Physical Intervention': self.physical_intervention_answer,
                                         'Restraint Used': self.restraint_used_answer,
                                         'Police Involved': self.police_involved_answer,
                                         'Requested By': self.requested_by_entry.get().strip(),
                                         'Contact Information': self.contact_information_entry.get().strip(),
                                         'Notes': self.notes_textbox.get('1.0', 'end-1c'),
                                         'Time Over 24 Hours': self.time_over_24_hours_answer
                                         }

        if ((self.row_to_append_saves_test['Date'] == '') and (self.row_to_append_saves_test['Call Received Time'] == '') and
            (self.row_to_append_saves_test['Arrival Time'] == '') and (self.row_to_append_saves_test['Completion Time'] == '') and
            (self.row_to_append_saves_test['Service Call Type'] == '') and (self.row_to_append_saves_test['Requested By'] == '') and
                (self.row_to_append_saves_test['Contact Information'] == '') and (self.row_to_append_saves_test['Notes'] == '')):

            tk.messagebox.showinfo(
                'Nothing to Save', 'To save a draft, you must have entered data.')
            return True

        else:

            self.identifier_answer = simpledialog.askstring(
                'Identifying Information', 'Please enter a Patient ID or patient name:')

            if self.identifier_answer is not None:

                self.row_to_append_saves = {
                    'Identifier': self.identifier_answer.strip(),
                    'Date': self.date_entry.get().strip(),
                    'Shift': self.var.get().strip(),
                    'Call Received Time': self.call_received_entry.get().strip(),
                    'Arrival Time': self.arrival_time_entry.get().strip(),
                    'Completion Time': self.completion_time_entry.get().strip(),
                    'Service Call Type': self.service_call_type_entry.get().strip(),
                    'Physical Intervention': self.physical_intervention_answer,
                    'Restraint Used': self.restraint_used_answer,
                    'Police Involved': self.police_involved_answer,
                    'Requested By': self.requested_by_entry.get().strip(),
                    'Contact Information': self.contact_information_entry.get().strip(),
                    'Notes': self.notes_textbox.get('1.0', 'end-1c'),
                    'Time Over 24 Hours': self.time_over_24_hours_answer
                }

                if self.saves_df.shape[0] == 10:
                    # Drop the oldest draft
                    self.saves_df.drop(0, axis=0, inplace=True)

                self.saves_df = self.saves_df.append(
                    self.row_to_append_saves, ignore_index=True)  # Use index to pull in later?

            else:
                return True

    def save_text_file(self):

        try:
            with open(PATH_LOGS + str(datetime.now())[0:19].replace(':', '-') + ' Incident Report Entry.txt', 'w+') as file:

                file.write('\n  Below is the information for the record submitted on: ' +
                           str(datetime.now())[0:19] + '\n')

                file.write('\n  Date: '.ljust(28) + self.row_to_append['Date'])
                file.write('\n  Shift: '.ljust(28) +
                           self.row_to_append['Shift'])
                file.write('\n  Call Received Time: '.ljust(28) +
                           self.row_to_append['Call Received Time'])
                file.write('\n  Arrival Time: '.ljust(28) +
                           self.row_to_append['Arrival Time'])
                file.write('\n  Completion Time: '.ljust(28) +
                           self.row_to_append['Completion Time'])
                file.write('\n  Service Call Type: '.ljust(28) +
                           self.row_to_append['Service Call Type'])
                file.write('\n  Physical Intervention: '.ljust(
                    28) + self.row_to_append['Physical Intervention'])
                file.write('\n  Restraint Used: '.ljust(28) +
                           self.row_to_append['Restraint Used'])
                file.write('\n  Police Involved: '.ljust(28) +
                           self.row_to_append['Police Involved'])
                file.write('\n  Requested By: '.ljust(28) +
                           self.row_to_append['Requested By'])
                file.write('\n  Contact Information: '.ljust(28) +
                           self.row_to_append['Contact Information'])
                file.write('\n\n  Notes:'.ljust(28) + '\n\n   ' +
                           self.row_to_append['Notes'].replace('\n', '\n   '))
        except:
            pass

    def clean_text_file_folder(self):

        FILES = [f for f in listdir(PATH_LOGS) if isfile(join(PATH_LOGS, f))]

        total_files = len(FILES)

        if total_files > MAX_TEXT_FILES:
            number_to_delete = (total_files - MAX_TEXT_FILES) + 30

            for count in range(number_to_delete):
                os.unlink(PATH_LOGS + '\\' + FILES[count])

    def save_files(self):
        '''FOR BOTH THE MONTHLY AND MASTER DATAFRAMES:
           Create a workbook, select the 1st worksheet, and title it. Convert dataframe to format for OpenPyXL, and set
           the column widths in advance. Loop through the new dataframe and insert the values into the Excel file, also specifying
           alignment and to wrap text. Save the Excel file.'''

        workbook = Workbook()
        worksheet = workbook.worksheets[0]

        rows = dataframe_to_rows(self.df, index=False)

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 17
        worksheet.column_dimensions['C'].width = 16
        worksheet.column_dimensions['D'].width = 21
        worksheet.column_dimensions['E'].width = 16
        worksheet.column_dimensions['F'].width = 21
        worksheet.column_dimensions['G'].width = 40
        worksheet.column_dimensions['H'].width = 24
        worksheet.column_dimensions['I'].width = 18
        worksheet.column_dimensions['J'].width = 19
        worksheet.column_dimensions['K'].width = 25
        worksheet.column_dimensions['L'].width = 25
        worksheet.column_dimensions['M'].width = 44
        worksheet.column_dimensions['N'].width = 30
        worksheet.column_dimensions['O'].width = 43
        worksheet.column_dimensions['P'].width = 43
        worksheet.column_dimensions['Q'].width = 30
        worksheet.column_dimensions['R'].width = 43
        worksheet.column_dimensions['S'].width = 45

        for row_index, row in enumerate(rows, 1):
            for column_index, value in enumerate(row, 1):
                worksheet.cell(row=row_index, column=column_index, value=value).alignment = Alignment(
                    horizontal='center', vertical='center', wrapText=True)

        workbook.save(PATH_MONTHLY + CURRENT_YEAR + '\\' + MONTH_ENTERED +
                      '\\Incident Reports - ' + CURRENT_YEAR + ' ' + MONTH_ENTERED[5:] + '.xlsx')

        # MASTER FILE ========================================================================================================
        workbook_master = Workbook()
        worksheet_master = workbook_master.worksheets[0]

        master_rows = dataframe_to_rows(self.master_df, index=False)

        worksheet_master.column_dimensions['A'].width = 15
        worksheet_master.column_dimensions['B'].width = 17
        worksheet_master.column_dimensions['C'].width = 16
        worksheet_master.column_dimensions['D'].width = 21
        worksheet_master.column_dimensions['E'].width = 16
        worksheet_master.column_dimensions['F'].width = 21
        worksheet_master.column_dimensions['G'].width = 40
        worksheet_master.column_dimensions['H'].width = 24
        worksheet_master.column_dimensions['I'].width = 18
        worksheet_master.column_dimensions['J'].width = 19
        worksheet_master.column_dimensions['K'].width = 25
        worksheet_master.column_dimensions['L'].width = 25
        worksheet_master.column_dimensions['M'].width = 44
        worksheet_master.column_dimensions['N'].width = 30
        worksheet_master.column_dimensions['O'].width = 43
        worksheet_master.column_dimensions['P'].width = 43
        worksheet_master.column_dimensions['Q'].width = 30
        worksheet_master.column_dimensions['R'].width = 43
        worksheet_master.column_dimensions['S'].width = 45

        for row_index, row in enumerate(master_rows, 1):
            for column_index, value in enumerate(row, 1):
                worksheet_master.cell(row=row_index, column=column_index, value=value).alignment = Alignment(
                    horizontal='center', vertical='center', wrapText=True)

        workbook_master.save(PATH_MASTER)

        try:
            workbook_master.save(PATH_MASTER_COPY)
        except:
            pass

    def save_drafts_file(self):
        '''Create a workbook, select the 1st worksheet, and title it. Convert drafts dataframe to format for OpenPyXL, and set
           the column widths in advance. Loop through the new dataframe and insert the values into the Excel file, also specifying
           alignment and to wrap text. Save the Excel file.'''

        workbook = Workbook()
        worksheet = workbook.worksheets[0]
        worksheet.title = 'Drafts'

        rows = dataframe_to_rows(self.saves_df, index=False)

        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 16
        worksheet.column_dimensions['D'].width = 21
        worksheet.column_dimensions['E'].width = 16
        worksheet.column_dimensions['F'].width = 21
        worksheet.column_dimensions['G'].width = 40
        worksheet.column_dimensions['H'].width = 24
        worksheet.column_dimensions['I'].width = 18
        worksheet.column_dimensions['J'].width = 19
        worksheet.column_dimensions['K'].width = 25
        worksheet.column_dimensions['L'].width = 23
        worksheet.column_dimensions['M'].width = 44
        worksheet.column_dimensions['N'].width = 30

        for row_index, row in enumerate(rows, 1):
            for column_index, value in enumerate(row, 1):
                worksheet.cell(row=row_index, column=column_index, value=value).alignment = Alignment(
                    horizontal='center', vertical='center', wrapText=True)

        workbook.save(PATH_DRAFTS)

    def reset_radio_buttons(self):

        self.var.set('7:30 - 19:30')

    def reset_checkboxes(self):
        '''Uncheck all Checkboxes.'''

        self.physical_intervention_checkbox.state(['!selected'])
        self.restraint_used_checkbox.state(['!selected'])
        self.police_involved_checkbox.state(['!selected'])
        self.time_over_24_hours_checkbox.state(['!selected'])

    def reset_entries(self):
        '''Reset all widgets, and also reset their colors.'''

        self.date_entry.delete(0, 'end')
        self.date_entry.config({'background': 'White'})

        self.call_received_entry.delete(0, 'end')
        self.call_received_entry.config({'background': 'White'})

        self.arrival_time_entry.delete(0, 'end')
        self.arrival_time_entry.config({'background': 'White'})

        self.completion_time_entry.delete(0, 'end')
        self.completion_time_entry.config({'background': 'White'})

        self.service_call_type_entry.delete(0, 'end')
        self.service_call_type_entry.config({'background': 'White'})

        self.requested_by_entry.delete(0, 'end')

        self.contact_information_entry.delete(0, 'end')

        self.notes_textbox.delete('1.0', 'end')

    def on_submit_button(self):
        '''Run validation functions on the Entry fields. If none have an error, get both the monthly and master dataframe, and the
           checkbox answers, and run a function to append the currently entered values to each dataframe. Save both dataframes as
           Excel files. Reset all the widgets. If the record submitted was an imported draft, delete that draft and update the
           Excel file. Reset the draft index to None as was initialized. Draft index is only not None when Select button
           function runs. Show a submission confirmation.'''

        self.date_validation()
        self.call_received_validation()
        self.arrival_time_validation()
        self.completion_time_validation()
        self.service_call_type_validation()

        if sum(self.errors.values()) == 0:

            self.get_dataframe()
            self.get_master_dataframe()
            self.get_checkbox_answers()
            self.append_row_to_df()
            self.save_files()
            self.save_text_file()
            self.clean_text_file_folder()
            self.reset_radio_buttons()
            self.reset_checkboxes()
            self.reset_entries()

            if self.draft_opened_index is not None:
                # Drop the draft from its dataframe
                self.saves_df.drop(self.draft_opened_index,
                                   axis=0, inplace=True)
                self.save_drafts_file()
                self.draft_opened_index = None  # Reset so that nothing is deleted on next save

            tk.messagebox.showinfo('Success', 'Entry Successfully Submitted!')

    def on_save_button(self):
        '''Get the drafts dataframe, and the values of the 4 Checkboxes. Run a function that updates the dataframe with current entered
           values if they are not all blank, otherwise displays a Messagebox and returns True. Only if the former occurs, the updated
           dataframe will be saved, and the program will terminate.'''

        self.get_saves_dataframe()
        self.get_checkbox_answers()  # Normally only called on Submit button
        check_if_empty = self.append_row_to_saves_df()

        if check_if_empty != True:
            self.save_drafts_file()
            self.destroy()  # Close the app

    def handle_topbox_listbox_creation(self):
        '''Create a header string to display above the listbox, use it as text to a Label, create a Listbox, get the drafts dataframe,
           and for each row, generate a fixed-width string of the values. Insert that string into the Listbox.'''

        header_string = '      ' + \
            'IDENTIFIER'.ljust(15) + 'DATE'.ljust(9) + \
            'SHIFT'.ljust(15) + 'CALL RECEIVED'.ljust(15)
        header_string += 'ARRIVAL'.ljust(10) + 'SERVICE CALL TYPE'.ljust(20)
        header_string += 'PHYS. INTERV.'.ljust(
            17) + 'RESTRAINT USED'.ljust(17) + 'POLICE INVOLVED'.ljust(17)
        header_string += 'REQUESTED BY'.ljust(15) + 'CONTACT INFO.'

        tk.Label(self.window, font=('Courier', 8), fg='blue',
                 text=header_string).grid(row=1, padx=(0, 4))

        self.top_lbox = tk.Listbox(self.window, font=(
            'Courier', 8), width=170, height=10, activestyle='none')
        self.top_lbox.grid(row=2, padx=(3, 0))

        self.get_saves_dataframe()

        index_number = 1

        for row in range(self.saves_df.shape[0]):
            row_string = ('(' + str(index_number) + ')').rjust(4)
            index_number += 1
            # Identifying Information
            row_string += '  ' + str(self.saves_df.iloc[row, 0])[:12].ljust(15)
            # Date
            row_string += str(self.saves_df.iloc[row, 1]
                              )[:5].replace('nan', '').ljust(9)
            row_string += self.saves_df.iloc[row, 2].ljust(15)  # Shift
            # Call Received Time          USING STR() PREVENTS INTEGERS (11:00) FROM SHOWING AS FLOATS WITH DECIMALS IN THE LISTBOX
            row_string += str(self.saves_df.iloc[row, 3]
                              )[:10].replace('nan', '').ljust(15)
            # Arrival Time
            row_string += str(self.saves_df.iloc[row, 4]
                              )[:7].replace('nan', '').ljust(10)
            # Service Call Type
            row_string += str(self.saves_df.iloc[row, 6]
                              )[:17].replace('nan', '').ljust(20)
            # Physical Intervention
            row_string += self.saves_df.iloc[row, 7].ljust(17)
            # Restraint Used
            row_string += self.saves_df.iloc[row, 8].ljust(17)
            # Police Involved
            row_string += self.saves_df.iloc[row, 9].ljust(17)
            # Requested By
            row_string += str(self.saves_df.iloc[row, 10]
                              )[:10].replace('nan', '').ljust(15)
            # Contact Information
            row_string += str(self.saves_df.iloc[row, 11]
                              )[:13].replace('nan', '')
#             row_string+= str(self.saves_df.iloc[row,11]).replace('\n', '     ').replace('nan', '')[:50].ljust(55) # Notes
#             row_string+= self.saves_df.iloc[row,12].ljust(8) # Time Over 24 Hours

            self.top_lbox.insert('end', row_string)

    def load_selected_draft(self):
        '''Set all widgets according the the values of the draft list.'''

        self.var.set(self.row_to_insert[2])  # Radio button

        if self.row_to_insert[7] == 'Yes':
            self.physical_intervention_checkbox.state(['selected'])
        if self.row_to_insert[8] == 'Yes':
            self.restraint_used_checkbox.state(['selected'])
        if self.row_to_insert[9] == 'Yes':
            self.police_involved_checkbox.state(['selected'])
        if self.row_to_insert[13] == 'Yes':
            self.time_over_24_hours_checkbox.state(['selected'])

        self.row_to_insert = [str(x).replace('nan', '')
                              for x in self.row_to_insert]

        self.date_entry.insert('end', str(self.row_to_insert[1]))
        self.call_received_entry.insert('end', str(
            self.row_to_insert[3]).replace('.0', ''))
        self.arrival_time_entry.insert('end', str(
            self.row_to_insert[4]).replace('.0', ''))
        self.completion_time_entry.insert('end', str(
            self.row_to_insert[5]).replace('.0', ''))
        self.service_call_type_entry.insert('end', str(self.row_to_insert[6]))
        self.requested_by_entry.insert('end', str(self.row_to_insert[10]))
        self.contact_information_entry.insert(
            'end', str(self.row_to_insert[11]).replace('.0', ''))
        self.notes_textbox.insert('end', str(self.row_to_insert[12]))

        self.window.destroy()

    def on_select_button(self):
        '''If no row is selected, throw an error. Otherwise, get the selected row, use it to extract the dataframe row as a list,
           store the index of the draft so it can be removed when submitted, reset the widgets, and fill with the extracted row'''

        try:
            row_selected = self.top_lbox.curselection()[0]
            self.row_to_insert = list(self.saves_df.iloc[row_selected, :])

            # If submitted, this index will locate the draft to delete
            self.draft_opened_index = row_selected

            self.reset_radio_buttons()
            self.reset_checkboxes()
            self.reset_entries()
            self.load_selected_draft()

        except:
            tk.messagebox.showinfo(
                'Selection Error', 'Please select a draft.', parent=self.window)

    def on_load_button(self):
        '''Try to load in the Draft Entries file. If no file, OR file is empty, FAIL. Otherwise, create a Toplevel window,
           place the Select button, and run a function to create the listbox.'''

        try:
            temp_df = pd.read_excel(PATH_DRAFTS)

            if temp_df.shape[0] == 0:
                fail = 1/0

            if self.window is None or not self.window.winfo_exists():
                self.window = tk.Toplevel()
                self.window.wm_title('Saved Drafts')
                self.window.wm_geometry('1200x224')
                self.window.resizable(False, False)

                self.select_button = tk.Button(self.window, font=(
                    'Arial', 11, 'bold'), fg='blue', text='CLICK HERE to Import Selected Draft', command=self.on_select_button)
                self.select_button.grid(
                    row=0, padx=(3, 0), pady=10, sticky='we')

                self.handle_topbox_listbox_creation()

        except:
            tk.messagebox.showinfo('No Drafts', 'There are no saved drafts.')


app = App()
app.mainloop()
